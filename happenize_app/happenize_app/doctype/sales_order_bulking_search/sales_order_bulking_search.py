# Copyright (c) 2023, Ahmed Abd El-Sattar and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe import publish_progress
import openpyxl as xl

class SalesOrderBulkingSearch(Document):
	@frappe.whitelist()
	def get_sales_orders(self):
		source = self.excel_sheet
		filename = source.split("/")
		file_path = frappe.get_site_path('private', 'files', filename[-1])
		work_book = xl.load_workbook(file_path)
		sheet = work_book[work_book.sheetnames[0]]
		total_rows = sheet.max_row - 1
		progress = 100/ total_rows
		start = 1
		for row in range(2,sheet.max_row+1):
			publish_progress(progress*start,title="Get Sales Orders...",description = f"""{start} Of {total_rows}""")
			start += 1
			sales_order_name = sheet.cell(row,1).value
			if frappe.db.exists('Sales Order', sales_order_name):
				self.append("sales_order_details", {
				"sales_order": sales_order_name,
			})
		self.save()
		self.reload()


